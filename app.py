import pandas as pd
from pandas import ExcelWriter, DataFrame
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import os
from dotenv import load_dotenv
import locale

# set locale
locale.setlocale(locale.LC_ALL, "pt_BR.UTF-8")

# load env variables
load_dotenv()
FILE_PATH = os.getenv("SALES_FILE_PATH")
SHEET_NAME = os.getenv("SHEET_NAME")



def Adjust_Cell_Size(df: DataFrame, worksheet: str):
    """Method to adjustment cell size to better reading in excel file.

    Args:
        df (DataFrame): DataFrame with data of sheet
        worksheet (str): Name of the sheet where DataFrame is located.
    """
    
    # Set all columns to a specific size based on content/value
    for col_idx, col in enumerate(df.columns, 1):
        max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2 # Get length
        col_letter = get_column_letter(col_idx) # Get col letter based on col index
        worksheet.column_dimensions[col_letter].width = max_length # Set size

def Format_Currency(df: DataFrame, cols: list):
    """Method to format all currency cols to the format R$ 999.999,99.

    Args:
        df (DataFrame): Dataframe where has a currency(ies) col
        col (str|list): Name(s) of the currency col(s)
    """
    
    # Set all currency cols to currency format R$ 999.999,99
    for col in cols:
        df[col] = df[col].apply(lambda x: locale.currency(x, symbol="R$", grouping=True)) # Format float value R$ 999.999,99



def Create_General_Report_Sheet(df_sales: DataFrame, writer: ExcelWriter):
    """Create General Report sheet with your specific columns and values.

    Args:
        df_sales (DataFrame): DataFrame with all sales.
        writer (ExcelWriter): ExcelWriter to the new excel file (Final Report)
    """
    
    # Data of the sheet
    data = {
        "Total de Vendas (R$)": float(df_sales["Valor Total"].sum()),
        "Total de Produtos Vendidos": int(df_sales["Quantidade Vendida"].sum()),
        "Vendedor com mais Vendas": df_sales[["Vendedor", "Quantidade Vendida"]].groupby("Vendedor").sum().idxmax().values[0],
        "Produto mais Vendido": df_sales[["Produto", "Quantidade Vendida"]].groupby("Produto").sum().idxmax().values[0],
        "Loja com mais Vendas": df_sales[["Localização", "Quantidade Vendida"]].groupby("Localização").sum().idxmax().values[0],
        "Vendas no Período": f"{df_sales["Data"].min()} - {df_sales["Data"].max()}"
    }
    
    new_df = pd.DataFrame([data]) # Create DataFrame
    Format_Currency(df=new_df, cols=["Total de Vendas (R$)"]) # Format currency cols
    
    new_df.to_excel(writer, sheet_name="Relatório Geral", index=False) # Transform the DataFrame to excel file in Writer
    Adjust_Cell_Size(df=new_df, worksheet=writer.sheets["Relatório Geral"]) # Adjustment cells size

def Create_Daily_Sales_Sheet(df_sales: DataFrame, writer: ExcelWriter):
    """Method to create sheet with sales by day.

    Args:
        df_sales (DataFrame): DataFrame with all sales.
        writer (ExcelWriter): ExcelWriter to manage the new excel file (create new sheet).
    """
    
    new_df = df_sales[["Data", "Quantidade Vendida", "Valor Total"]].groupby("Data").sum() # Create new DataFrame with sales group by day
    Format_Currency(df=new_df, cols=["Valor Total"]) # Format currency col
    
    new_df.to_excel(writer, sheet_name="Vendas Diárias", index=True) # Transform the DataFrame to excel file in Writer
    Adjust_Cell_Size(df=new_df, worksheet=writer.sheets["Vendas Diárias"]) # Adjustment cells size

def Create_Sales_by_Product_Sheet(df_sales: DataFrame, writer: ExcelWriter):
    """Method to create sheet with sales by product.

    Args:
        df_sales (DataFrame): DataFrame with all sales.
        writer (ExcelWriter): ExcelWriter to manage the new excel file (create new sheet).
    """
    
    new_df = df_sales[["Produto", "Quantidade Vendida", "Valor Total"]].groupby("Produto").sum() # Create new DataFrame with sales group by product
    Format_Currency(df=new_df, cols=["Valor Total"]) # Format currency col
    
    new_df.to_excel(writer, sheet_name="Vendas por Produto", index=True) # Transform the DataFrame to excel file in Writer
    Adjust_Cell_Size(df=new_df, worksheet=writer.sheets["Vendas por Produto"]) # Adjustment cells size

def Create_Sales_by_Seller_Sheet(df_sales: DataFrame, writer: ExcelWriter):
    """Method to create sheet with sales by seller.

    Args:
        df_sales (DataFrame): DataFrame with all sales.
        writer (ExcelWriter): ExcelWriter to manage the new excel file (create new sheet).
    """
    
    new_df = df_sales[["Vendedor", "Quantidade Vendida", "Valor Total"]].groupby("Vendedor").sum() # Create new DataFrame with sales group by seller
    Format_Currency(df=new_df, cols=["Valor Total"]) # Format currency col
    
    new_df.to_excel(writer, sheet_name="Vendas por Vendedor", index=True) # Transform the DataFrame to excel file in Writer
    Adjust_Cell_Size(df=new_df, worksheet=writer.sheets["Vendas por Vendedor"]) # Adjustment cells size

def Create_Sales_by_Location_Sheet(df_sales: DataFrame, writer: ExcelWriter):
    """Method to create sheet with sales by location.

    Args:
        df_sales (DataFrame): DataFrame with all sales.
        writer (ExcelWriter): ExcelWriter to manage the new excel file (create new sheet).
    """
    
    new_df = df_sales[["Localização", "Quantidade Vendida", "Valor Total"]].groupby("Localização").sum() # Create new DataFrame with sales group by location
    Format_Currency(df=new_df, cols=["Valor Total"]) # Format currency col
    
    new_df.to_excel(writer, sheet_name="Vendas por Localização", index=True) # Transform the DataFrame to excel file in Writer
    Adjust_Cell_Size(df=new_df, worksheet=writer.sheets["Vendas por Localização"]) # Adjustment cells size



def Create_Total_Value_by_Product_Chart(df_sales: DataFrame, ws):
    """Method to create a bar chart with the total value by product

    Args:
        df_sales (DataFrame): Dataframe with all sales.
        ws (sheet): Chart sheet of the output excel file.
    """
    
    df_sales_by_product = df_sales[["Produto", "Valor Total"]].groupby("Produto").sum()
    
    # create and configure bar chart
    fig, ax = plt.subplots(figsize=(15, 6))
    df_sales_by_product.plot(kind="bar", ax=ax)
    ax.set_title("Valor Total de Vendas por Produto")
    ax.grid(axis='y', color='r', linestyle='--')

    ax.set_xticklabels(df_sales_by_product.index, rotation=0)

    ax.set_ylabel("Valor Total (R$)")
    ax.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x, _: f'R$ {x:,.2f}'.replace(",", "X").replace(".", ",").replace("X", "."))) # Transform y axis values to original prices values and format to R$ 999.999,99
    
    plt.legend(bbox_to_anchor=(-0.006, 1.08), loc='upper left')
    
    fig.savefig("graphics/total_value_by_product_graphic.png", format='png', dpi=300, bbox_inches='tight')
    
    
    # Add chart in excel file
    img = Image("graphics/total_value_by_product_graphic.png")
    ws.add_image(img, "B2")
    img.width = 800
    img.height = 400

def Create_Sales_by_Location_Chart(df_sales: DataFrame, ws):
    """Method to create a pie chart with the sales by location

    Args:
        df_sales (DataFrame): Dataframe with all sales.
        ws (sheet): Chart sheet of the output excel file.
    """
    
    df_sales_by_location = df_sales[["Localização", "Quantidade Vendida"]].groupby("Localização").sum()
    
    # Create and configure pie chart
    fig, ax = plt.subplots()
    df_sales_by_location.plot(kind="pie", y="Quantidade Vendida", ax=ax, autopct='%1.1f%%')

    plt.title("Distribuição das Vendas por Localização")
    plt.ylabel("")
    plt.legend(bbox_to_anchor=(0.070, 0.97))
    
    fig.savefig("graphics/sales_by_location_graphic.png", format='png', dpi=300, bbox_inches='tight')
    
    
    
    # add charset in excel file
    img = Image("graphics/sales_by_location_graphic.png")
    ws.add_image(img, "B26")
    img.width = 300
    img.height = 300

def Create_Evolution_Sales_Chart(df_sales: DataFrame, ws):
    """Method to create a line chart with the evolution of sales.

    Args:
        df_sales (DataFrame): Dataframe with all sales.
        ws (sheet): Chart sheet of the output excel file.
    """
    
    df = df_sales[["Data", "Quantidade Vendida"]].groupby("Data").sum()
    
    # Create and configure line chart
    fig, ax = plt.subplots(figsize=(15, 6))
    df.plot(kind='line', title='Evolução da Quantidade de Vendas', ax=ax)
    plt.ylabel("Quantidade Vendida")
    plt.grid(color='r', linestyle='--')
    
    fig.savefig("graphics/evolution_of_sales_graphic.png", format='png', dpi=300, bbox_inches='tight')
    
    
    
    # add chart in excel file
    img = Image("graphics/evolution_of_sales_graphic.png")
    ws.add_image(img, "B45")
    img.width = 850
    img.height = 350

def Create_Graphics_Sheet(df_sales: DataFrame, writer: ExcelWriter):
    """Method to create a sheet with all charts.

    Args:
        df_sales (DataFrame): Dataframe with all sales.
        writer (ExcelWriter): ExcelWriter variable.
    """
    
    # Create excel sheet with openpyxl
    wb = writer.book
    wb.create_sheet("Gráficos")
    ws = wb["Gráficos"]

    Create_Total_Value_by_Product_Chart(df_sales, ws)
    Create_Sales_by_Location_Chart(df_sales, ws)
    Create_Evolution_Sales_Chart(df_sales, ws)



def Create_General_Report_File(df_sales: DataFrame):
    """Method to create the output file.

    Args:
        df_sales (DataFrame): Dataframe with all sales.
    """
    with pd.ExcelWriter(path="output/General_Report.xlsx", date_format="%d/%m/%Y") as writer:
        
        Create_General_Report_Sheet(df_sales, writer)
        writer.book.save("output/General_Report.xlsx")
        
        Create_Daily_Sales_Sheet(df_sales, writer)
        Create_Sales_by_Product_Sheet(df_sales, writer)
        Create_Sales_by_Seller_Sheet(df_sales, writer)
        Create_Sales_by_Location_Sheet(df_sales, writer)
        Create_Graphics_Sheet(df_sales, writer)

def Read_Excel():
    """Read excel file and return a DataFrame.

    Returns:
        DataFrame: DataFrame with the data of the excel file.
    """
    
    return pd.read_excel(FILE_PATH, SHEET_NAME)

if __name__ == "__main__":
    # read excel
    df_sales = Read_Excel()
    
    # Create General Report
    Create_General_Report_File(df_sales)